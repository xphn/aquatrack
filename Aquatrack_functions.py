# --------------------
# This program reads a .csv file containing WeatherUnderground station names, start date, and end date.
# It gathers rain data from Wunderground and writes it to .csv files for each station name.
#
from bs4 import BeautifulSoup
import datetime
import dateparser
import re
import requests
import pandas as pd
import simplekml
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
from selenium import webdriver
import time
import os

# Function to convert string to datetime format
def convert(date_time_in):
        datetime_str = dateparser.parse(date_time_in)
        return datetime_str


# Function to calculate date range
def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + datetime.timedelta(n)


# Function to collect data for one day
def fetch_one_day(station, date, paccumchoice, savefolder):
    # # debug line
    # station = 'Kcaburli4'
    # date = endDate.strftime("%Y-%m-%d")
    paccumchoice = 'yes'

    url = 'https://www.wunderground.com/dashboard/pws/' + station.upper() + '/table/' + date + '/' + date + '/daily'
    print('fetching page', url)
    page = requests.get(url)

    soup = BeautifulSoup(page.content, 'html.parser')
    # rows = soup.select('.history-table tr')

    table_heads = soup.select('table.desktop-table.history-table thead tr th')
    table_rows = soup.select('table.desktop-table.history-table tbody tr')

    data_all_rows = []
    for row in table_rows:
        data_row = []
        all_cells = row.select("td")
        for cell in all_cells:
            try:
                data_row.append(cell.text.replace(u'\xa0°', u' ').strip())
            except:
                data_row.append(np.nan)
        data_all_rows.append(data_row)

    head_names = []
    for head in table_heads:
        head_names.append(head.text)

    df_data = pd.DataFrame(data_all_rows)
    df_data.columns = head_names

    df_data_export = df_data.copy()
    df_data_export['Precip. Accum.'] = df_data_export['Precip. Accum.'].str.extract('(\d*\.\d+|\d+)',expand=False).astype(float)
    df_data_export['Precip. Rate.'] = df_data_export['Precip. Rate.'].str.extract('(\d*\.\d+|\d+)',expand=False).astype(float)
    df_data_export['Time'] = date + ' '+ df_data_export['Time']
    df_data_export = df_data_export[['Time', 'Precip. Rate.', 'Precip. Accum.']].replace('--', np.NAN)
    df_data_export = df_data_export.rename(columns = {'Time':"datetime", "Precip. Rate.":"prate", 'Precip. Accum.':"paccum"})
    df_data_export.reset_index(drop=True, inplace=True)
    # df_data_export.to_csv(savefolder + '/'+station + '.csv', index=False)

    return df_data_export



# Function to Collect Rain Data for one station, every day from start date to end date
def collect_all_days(station, start_date, end_date, paccumchoice, savefolder):
    '''
    This function obtains all RG data from the website and return a list of rain gauge names that are not available
    on the webiste

    :param station: station name from the list
    :param start_date: start date from the list
    :param end_date: end date from the list
    :param paccumchoice: may need to used in the future
    :param savefolder: export csv files to
    :return: a list of rain gauge names that are not available
    on the webiste
    '''
    dfs = []
    NA_station = []
    for single_date in daterange(start_date, end_date):
        try:
            df_single = fetch_one_day(station, single_date.strftime("%Y-%m-%d"), paccumchoice,savefolder)
        # df_single.to_csv(savefolder + '/' + station +'.csv', index=False)
            dfs.append(df_single)
        except KeyError:
            print(f'Oops, The data for Station {station} is not available at {single_date}, please check the website and consider changing the '
                  f'date range or just skipping this station.')

        try:
            df_alldays = pd.concat(dfs)
            df_alldays.to_csv(savefolder + '/'+station + '.csv', index=False)
        except ValueError:
            print(f'No data has been collected for {station}')
            NA_station.append(station)
    return NA_station

def coordinate (station,date):
    url = 'https://www.wunderground.com/dashboard/pws/' + station.upper() + '/table/' + date + '/' + date + '/daily'
    print('fetching coordinate')
    page = requests.get(url)

    soup = BeautifulSoup(page.content, 'html.parser')
    #rows = soup.select('.history-table tr')

    ## finding hidden longitude and latitude
    try:
        test = soup.find_all("script", attrs={'id': 'app-root-state'})
        test_content = test[0].contents[0]

        pattern_lon = re.compile(r"lon&q;:(.*?),&q;")
        pattern_lat = re.compile(r"lat&q;:(.*?),&q;")
        lon = (pattern_lon.findall(test_content)[0])
        lat = (pattern_lat.findall(test_content)[0])
        print(f'The longitude value for Station {station} is: {lon}')
        print(f'The latitude value for Station {station} is: {lat}')
    except:  #if the backend returns nothing, then mannually open the web page and click the detail button
        path = "C:\Program Files (x86)\chromedriver.exe"
        driver = webdriver.Chrome(path)
        driver.get('url')

        time.sleep(2)
        button = driver.find_element_by_xpath(
            '//*[@id="inner-content"]/div[1]/app-dashboard-header/div[2]/div/div[2]/div/lib-pws-info-icon/mat-icon')
        button.click()

        divs = driver.find_element_by_class_name('cdk-overlay-container')

        a = divs.text
        b = list(filter(lambda x: 'Latitude / Longitude' in x, a.split('\n')))[0]

        Num = re.findall(r'[0-9]+', b)
        lat = int(Num[0]) + int(Num[1]) * 0.001
        lon = -int(Num[2]) - int(Num[3]) * 0.001

        print(f'The longitude value for Station {station} is: {lon}')
        print(f'The latitude value for Station {station} is: {lat}')
        time.sleep(1)
        driver.quit()

    return lon, lat

# Making kml file
def kml_making(df_coordinate_all, savefolder, stationlist):
    kml = simplekml.Kml()
    for n in range(len(df_coordinate_all)):
        Name = df_coordinate_all.index[n]
        Coords = [(float(df_coordinate_all.loc[Name,'Longitude (Degree)']), float(df_coordinate_all.loc[Name,'Latitude (Degree)']))]
        kml.newpoint(name = Name, coords=Coords) # lon, lat, optional height
        print(Coords)
    kml.save(savefolder+'/'+os.path.split(stationlist)[1].split('.')[0]+'.kml')


def fill_excel(stationName,exceltemp, savefolder):
    wb = load_workbook(exceltemp)
    ws = wb.active
    #  import station data
    if os.path.isfile(savefolder+'/'+ stationName + '.csv'):
        a = pd.read_csv(savefolder+'/'+ stationName + '.csv', header=0)
        # Remove extra headers
        a = a.drop(a[a['datetime'] == 'datetime'].index)
        a['datetime'] = pd.to_datetime(a['datetime'], format='%Y-%m-%d %I:%M %p')
        a.reset_index(inplace=True)
        a.paccum = a.paccum.astype(float)
        a.drop(['index'], axis=1, inplace=True) # this is probably no longer needed as we have updated the webscraping code. Now we won't see bunch of 'datatime' rows
        #  Convert excel rows
        rows = dataframe_to_rows(a, index=False)
        # fill the excel table

        for r_idx, row in enumerate(rows, 6):  # starts at 6 as you want to skip the first 5 rows
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx>10:
                ws[f'D{r_idx}'] = Translator(ws['D10'].value, origin='D10').translate_formula(f'D{r_idx}')
                ws[f'E{r_idx}'] = Translator(ws['E10'].value, origin='E10').translate_formula(f'E{r_idx}')
                ws[f'F{r_idx}'] = Translator(ws['F10'].value, origin='F10').translate_formula(f'F{r_idx}')
                ws[f'H{r_idx}'] = Translator(ws['H10'].value, origin='H10').translate_formula(f'H{r_idx}')
                ws[f'I{r_idx}'] = Translator(ws['I10'].value, origin='I10').translate_formula(f'I{r_idx}')
                ws[f'J{r_idx}'] = Translator(ws['J10'].value, origin='J10').translate_formula(f'J{r_idx}')
                ws[f'K{r_idx}'] = Translator(ws['K10'].value, origin='K10').translate_formula(f'K{r_idx}')
                ws[f'L{r_idx}'] = Translator(ws['L10'].value, origin='L10').translate_formula(f'L{r_idx}')

        ws.auto_filter.ref = f'H7:L{ws.max_row}'
        ws.auto_filter.add_filter_column(0, ['Keep'], blank = False)
        ws.auto_filter.add_sort_condition(f'H7:H{ws.max_row}')
        wb.save(savefolder+'/'+ stationName+"_processed.xlsx")



def get_idf():
    coordinate_list = tkinter.filedialog.askopenfilename(
        title='Select the coordinate list file')
    coordinate_pd = pd.read_csv(coordinate_list)

    saveto = os.path.dirname(coordinate_list)

    for i in range(len(coordinate_pd)):
        station_name = coordinate_pd.iloc[i,0]
        lon = str(coordinate_pd.iloc[i,1])
        lat = str(coordinate_pd.iloc[i,2])
        url_csv = f"https://hdsc.nws.noaa.gov/cgi-bin/hdsc/new/fe_text_mean.csv?lat={lat}6&lon={lon}&data=depth&units=english&series=pds&selAddr=Burlingame, California, USA&selElevNum=511.04&selElevSym=ft&selStaName=-"
        req_csv = requests.get(url_csv)

        with open(f"{saveto}/{station_name}_idf.csv", 'w', encoding=req_csv.encoding) as csvFile:
            csvFile.write(req_csv.text, )



# Ask if want precipitation accumulation in addition to precipitation rate.
# If the station's precipitation rate ="--", even if choose no, would default and output precipitation accumulation
# #paccumchoice = input("Do you want precipitation accumulation in addition to precipitation rate? Yes/No: ")
# # paccumchoice = "Yes"
# #
# # # Accessing a text file - www.101computing.net/mp3-playlist/
# # stationlist = input("Enter exact csv filename of station list, E.g. stationlist.csv, : ")
# # stationlist = 'stationlist_paccum.csv'
# file = open(stationlist, "r")
# all_coordinate_dict = {}
# # Repeat for each station in the .csv file
# for line in file:
#     # Let's split the line into an array called "fields" using the "," as a separator:
#     fields = line.split(",")
#
#     # and let's extract the data:
#     stationName = fields[0]
#     startDate_str = fields[1]
#     endDate_str = fields[2]
#
#     startDate = convert(startDate_str)
#     endDate = convert(endDate_str)
#     print("Get " + stationName + " from: " + startDate_str + " to: " + endDate_str)
#
#     # collect rain data and the rain gauge coordination
#     collect_all_days(stationName, startDate, endDate, paccumchoice)
#     coordinates = coordinate(stationName, startDate.strftime("%Y-%m-%d"))
#     all_coordinate_dict[stationName] = coordinates
#
#     df_coordinate_all = pd.DataFrame(all_coordinate_dict).T.rename(columns={1:'Latitude (Degree)', 0:'Longitude (Degree)'})
#     fill_excel(stationName)
#
#
#
# # It is good practice to close the file at the end to free up resources
# file.close()
#
# # save the coordination file
# df_coordinate_all.to_csv(f'Coordination for {stationlist}')
#
# # Making KML file
# kml_making(df_coordinate_all)
































#######Testing of extracting lon and lat
# url = " https://www.wunderground.com/dashboard/pws/KCASANRA706/table/2020-01-15/2020-01-15/daily"
# page = requests.get(url)
# soup = BeautifulSoup(page.content, 'html.parser')
#
# test = soup.find_all("script", attrs={'id': 'app-root-state'})
# test_content = test[0].contents[0]
#
# test_content.find('longitude')
#
# import re
# import requests
# pattern_lon = re.compile(r"lon&q;:(.*?),&q;")
# pattern_lat = re.compile(r"lat&q;:(.*?),&q;")
# print(pattern_lon.findall(test_content)[0])
# print(pattern_lat.findall(test_content)[0])